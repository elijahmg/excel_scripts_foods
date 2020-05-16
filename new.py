import openpyxl as excel
import json

from new_excel import NewExcel


class NewSheet:
    __money_url = ''
    money_class = None
    money_workbook = None
    money_sheet = None

    new_excel_obj_eng = None
    new_excel_obj_cz = None
    first_column_names = []

    def __init__(self, money_url, radio_button):
        self.__money_url = money_url
        self.money_workbook = excel.open(self.__money_url)
        money_sheet_names = self.money_workbook.sheetnames
        self.money_sheet = self.money_workbook[money_sheet_names[0]]

        self.new_excel_obj_eng = NewExcel()
        self.new_excel_obj_cz = NewExcel()
        self.radio_button = radio_button
        try:
            with open('config.json') as json_file:
                self.config = json.load(json_file)
        except Exception:
            raise KeyError

    def set_values(self, row_index, column_index, value, row):
        """ Set values to the cell"""

        eng_value = value
        cz_value = value

        # Get column with brand according to config
        brand_column_name = self.config['en_brand_name_column'] if self.radio_button\
            else self.config['cz_brand_name_column']

        # Get column with cz translate according to config
        cz_translate_column = self.config['en_config_cz_name_column'] if self.radio_button\
            else self.config['cz_config_cz_name_column']

        if column_index == self.config['columns_config']['name']:
            if row_index != 0:
                index_cz_translate = self.first_column_names.index(cz_translate_column)
                cz_value = row[index_cz_translate].value

                if cz_value is None or cz_value == '':
                    cz_value = eng_value
                brand_column_index = self.first_column_names.index(brand_column_name)
                brand_name = row[brand_column_index].value

                if brand_name is not None and brand_name != '':
                    if brand_name in eng_value:
                        eng_value = eng_value.replace(brand_name, '').strip()

                    eng_value = brand_name + ' ' + eng_value

                    if brand_name in cz_value:
                        cz_value = cz_value.replace(brand_name, '').strip()

                    cz_value = brand_name + ' ' + cz_value

        self.new_excel_obj_eng.sheet.cell(row=row_index + 1, column=column_index + 1, value=eng_value)
        self.new_excel_obj_cz.sheet.cell(row=row_index + 1, column=column_index + 1, value=cz_value)

    def create_new_sheet(self):
        """
        Create new sheets for e-shop

        Rules:
        code - column_index = 0
        pairCode - column_index = 1
        name - column_index = 2
        """

        custom_column_index = 3
        is_custom_column_filling = False
        config = self.config['en'] if self.radio_button\
            else self.config['cz']

        self.first_column_names = [cell[0].value for cell in self.money_sheet.columns]

        for column_index, column in enumerate(self.money_sheet.columns):
            column_value = column[0].value
            # Check if value is in dict
            if column_value in config:
                # Move custom column index
                if is_custom_column_filling:
                    custom_column_index = custom_column_index + 1

                for row_index, row in enumerate(self.money_sheet.rows):
                    # Get a value from dictionary, if not find
                    # in dictionary take value from excel sheet
                    dict_value = str(column[row_index].value)
                    value = config.get(dict_value, column[row_index].value)
                    config_column_index = self.config['columns_config'].get(config[column_value], 'custom')

                    new_sheet_column_index = custom_column_index if config_column_index == 'custom' \
                        else config_column_index

                    is_custom_column_filling = config_column_index == 'custom'

                    self.set_values(row_index=row_index, column_index=new_sheet_column_index, value=value, row=row)
            # else:
            #     for row_index, row in enumerate(self.money_sheet.rows):
            #         if row_index == 0:
            #             pair_code_column_index = columns_config['pairCode']
            #             self.set_values(row_index=row_index, column_index=pair_code_column_index, value='pairCode')
            #             break

        self.new_excel_obj_eng.workbook.save(self.__money_url.replace('.xlsx', '_for_eng_eshop.xlsx'))
        self.new_excel_obj_cz.workbook.save(self.__money_url.replace('.xlsx', '_for_cz_eshop.xlsx'))
