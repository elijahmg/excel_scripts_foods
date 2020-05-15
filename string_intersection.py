import math
import openpyxl as excel
from openpyxl.styles import PatternFill
from openpyxl.styles.colors import YELLOW
from difflib import SequenceMatcher
import re

import time

url_money = 'C://Users//ww//Desktop//excel//Artikl - 13.4.2020.xlsx'
url_eshop = 'C://Users//ww//Desktop//excel//Codes for migration(intersection 100).xlsx'

wb_money = excel.open(url_money)
ws_money = wb_money[wb_money.sheetnames[0]]

wb_eshop = excel.open(url_eshop)
ws_eshop = wb_eshop[wb_eshop.sheetnames[0]]


def remove_char(str):
    return re.sub(r'[\s-]+', '', str).lower()


def build_sheet():
    temp_code = ''
    intersection = 65
    url_money = 'C://Users//ww//Desktop//excel//eshp//full_migration//filtered_v3.xlsx'
    url_eshop = 'C://Users//ww//Desktop//excel//eshp//full_migration//products_17_4.xlsx'
    wb_money = excel.open(url_money)
    ws_money = wb_money[wb_money.sheetnames[0]]

    wb_eshop = excel.open(url_eshop)
    ws_eshop = wb_eshop[wb_eshop.sheetnames[0]]

    new_book = excel.Workbook()
    new_sheet = new_book.active

    # exited = 'C://Users//ww//Desktop//excel//eshp//full_migration//New codes (hope so)_v1.xlsx'
    # exited_book = excel.open(exited)
    # exited_sheet = exited_book[exited_book.sheetnames[0]]
    #
    # existed_code = [row[0].value for row in exited_sheet.rows]

    # while intersection > 85:
    #     print('intersection', intersection)
    #     new_sheet.append(['intersection', intersection])

    for index, row_eshop in enumerate(ws_eshop.rows):
        for row_money in ws_money.rows:
            name_eshop = remove_char(row_eshop[2].value)
            name_money = remove_char(row_money[2].value).split(';')[0]

            intersection_value = SequenceMatcher(a=name_eshop, b=name_money).ratio() * 100
            intersection_value = math.ceil(intersection_value)

            if intersection_value > intersection:
                # print('intersection_value', intersection_value)
                # print(row_eshop[2].value, '/------\\', row_money[2].value.split(';')[0])
                # print('=' * 30)
                whole_row = [cell.value for cell in row_eshop]
                whole_row[0] = row_money[0].value
                whole_row[3] = row_money[2].value
                new_sheet.append(whole_row)
        # intersection -= 1

    new_book.save('C://Users//ww//Desktop//excel//eshp//full_migration//New codes (hope so)_v3.xlsx')


def non_existed_prods():
    url_money = 'C://Users//ww//Desktop//excel//eshp//27_04//en_products_missing_cz.xlsx'
    url_eshop = 'C://Users//ww//Desktop//excel//eshp//27_04//cz_products_missing_cz.xlsx'
    wb_money = excel.open(url_money)
    ws_money = wb_money[wb_money.sheetnames[0]]

    wb_eshop = excel.open(url_eshop)
    ws_eshop = wb_eshop[wb_eshop.sheetnames[0]]

    new_book = excel.Workbook()
    new_sheet = new_book.active
    codes_money = [row_eshop[0].value for row_eshop in ws_eshop]

    for row_eshop in ws_money.rows:
        money_code = row_eshop[0].value
        if money_code not in codes_money:
            new_sheet.append([cell.value for cell in row_eshop])

    new_book.save('C://Users//ww//Desktop//excel//eshp//27_04//Not in cz eshop.xlsx')


def set_cat():
    url_money = 'C://Users//ww//Desktop//excel//eshp//3//full_products.xlsx'
    wb_money = excel.open(url_money)
    ws_money = wb_money[wb_money.sheetnames[0]]

    url_categ = 'C://Users//ww//Desktop//excel//eshp//3//cz_categories.xlsx'
    wb_categ = excel.open(url_categ)
    ws_categ = wb_categ[wb_categ.sheetnames[0]]

    for index, eshop_row in enumerate(ws_money.rows):
        def_cat = eshop_row[3].value
        cat_1 = eshop_row[4].value
        cat_2 = eshop_row[5].value
        cat_3 = eshop_row[6].value

        for cat_row in ws_categ.rows:
            eng_cat = cat_row[0].value
            cz_cat = cat_row[1].value

            if def_cat is not None and eng_cat in def_cat:
                eshop_row[3].value = eshop_row[3].value.replace(eng_cat, cz_cat)
            if cat_1 is not None and eng_cat in cat_1:
                eshop_row[4].value = eshop_row[4].value.replace(eng_cat, cz_cat)
            if cat_2 is not None and eng_cat in cat_2:
                eshop_row[5].value = eshop_row[5].value.replace(eng_cat, cz_cat)
            if cat_3 is not None and eng_cat in cat_3:
                eshop_row[6].value = eshop_row[6].value.replace(eng_cat, cz_cat)
    wb_money.save('C://Users//ww//Desktop//excel//eshp//3//en_prod_vis_23_4_translated.xlsx')


def change_names():
    url_money = 'C://Users//ww//Desktop//excel//eshp//3//Artikl_23_04.xlsx'
    wb_money = excel.open(url_money)
    ws_money = wb_money[wb_money.sheetnames[0]]

    url_eshop = 'C://Users//ww//Desktop//excel//eshp//3//en_prod_vis_23_4_translated.xlsx'
    wb_eshop = excel.open(url_eshop)
    ws_eshop = wb_eshop[wb_eshop.sheetnames[0]]

    for row_eshop in ws_eshop:
        for row_money in ws_money:
            code_money = row_money[2].value
            code_eshop = row_eshop[0].value

            if code_money == code_eshop and row_money[3].value:
                row_eshop[2].value = row_money[3].value
    wb_eshop.save('C://Users//ww//Desktop//excel//eshp//3//cz_prod_vis_23_4_to_import.xlsx')

def test():
    print('test')

def filter_all_uniq():
    url_money = 'C://Users//ww//Desktop//excel//Artikl - 13.4.2020 (to use).xlsx'
    url_eshop = 'C://Users//ww//Desktop//excel//eshp//v1//New codes (hope so).xlsx'

def clean_the_shit():
    url_money = 'C://Users//ww//Desktop//excel//eshp//24_04//Artikl_24_04.xlsx'
    wb_money = excel.open(url_money)
    ws_money = wb_money[wb_money.sheetnames[0]]

    url_eshop = 'C://Users//ww//Desktop//excel//eshp//3//en_products_back_up_24_04.xlsx'
    wb_eshop = excel.open(url_eshop)
    ws_eshop = wb_eshop[wb_eshop.sheetnames[0]]

    new_book = excel.Workbook()
    new_sheet = new_book.active

    array_money_codes = [row[4].value for row in ws_money.rows]

    for row_eshop in ws_eshop.rows:
        eshop_code = row_eshop[0].value

        if eshop_code in array_money_codes:
            new_sheet.append([cell.value for cell in row_eshop])
    new_book.save('C://Users//ww//Desktop//excel//eshp//3//cleaned_en_eshop.xlsx')


def prices():
    url_money = 'C://Users//ww//Desktop//excel//eshp//28_04//PolozkaCeniku_20_43.xlsx'
    wb_money = excel.open(url_money)
    ws_money = wb_money[wb_money.sheetnames[0]]

    url_eshop = 'C://Users//ww//Desktop//excel//eshp//28_04//en_products_prices_21_00.xlsx'
    wb_eshop = excel.open(url_eshop)
    ws_eshop = wb_eshop[wb_eshop.sheetnames[0]]

    for row_money in ws_money.rows:
        for index, row_eshop in enumerate(ws_eshop.rows):
            money_code = row_money[4].value
            eshop_code = row_eshop[0].value

            if eshop_code == money_code:
                ws_eshop.cell(row=index + 1, column=11, value=row_money[2].value)
                ws_eshop.cell(row=index + 1, column=12, value=row_money[10].value)
                if row_eshop[3].value != row_money[2].value:
                    for cell in row_eshop:
                        cell.fill = PatternFill('solid', fgColor=YELLOW)
                # row_eshop[9].value = row_money[9].value
                # row_eshop[10].value = row_money[10].value
                # 9 column index, code 3, price 9, 10
    wb_eshop.save('C://Users//ww//Desktop//excel//eshp//28_04//en_products_compare_21_00.xlsx')


start = time.perf_counter()
print('start', start)

# non_existed_prods()
# set_cat()
# change_names()
# clean_the_shit()
prices()

end = time.perf_counter()
print(end - start)

# test()
