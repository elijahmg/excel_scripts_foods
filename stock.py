import openpyxl as excel

stock_id = 8
av_out_stock = 9
av_in_stock = 10
unit = 11

out_stock = 'Out of Stock'
in_stock = 'In Stock'


def set_stock():
    url_money = 'C://Users//ww//Desktop//excel//eshp//24_04//Artikl_24_04_stock.xlsx'
    wb_money = excel.open(url_money)
    ws_money = wb_money[wb_money.sheetnames[0]]

    url_eshop = 'C://Users//ww//Desktop//excel//eshp//24_04_17_00//cz_stock.xlsx'
    wb_eshop = excel.open(url_eshop)
    ws_eshop = wb_eshop[wb_eshop.sheetnames[0]]

    for e_shop_row in ws_eshop.rows:
        for money_row in ws_money.rows:

            money_code = money_row[0].value
            e_eshop_code = e_shop_row[0].value

            if money_code == e_eshop_code:
                money_stock = money_row[1].value

                if money_stock > 0:
                    e_shop_row[6].value = 'blockUnregistered'
                    e_shop_row[3].value = in_stock
                    e_shop_row[4].value = in_stock
                else:
                    e_shop_row[6].value = 'hidden'
                    e_shop_row[3].value = out_stock
                    e_shop_row[4].value = out_stock

            # money_code = money_row[0].value
            # e_eshop_code = e_shop_row[0].value
            #
            # if money_code == e_eshop_code:
            #     money_stock = money_row[1].value
            #     money_unit = money_row[2].value
            #
            #     e_shop_unit = e_shop_row[unit].value
            #
            #     if e_shop_unit == money_unit:
            #         e_shop_row[stock_id].value = money_stock
            #         if money_stock == 0:
            #             e_shop_row[av_in_stock].value = out_stock
            #             e_shop_row[av_out_stock].value = out_stock
            #         else:
            #             e_shop_row[av_in_stock].value = in_stock
            #             e_shop_row[av_out_stock].value = in_stock
            #
            #     elif e_shop_unit == '100g' and money_unit == 'kg':
            #         e_shop_row[stock_id].value = money_stock * 10
            #
            #         if money_stock == 0:
            #             e_shop_row[av_in_stock].value = out_stock
            #             e_shop_row[av_out_stock].value = out_stock
            #         else:
            #             e_shop_row[av_in_stock].value = in_stock
            #             e_shop_row[av_out_stock].value = in_stock

    wb_eshop.save('C://Users//ww//Desktop//excel//eshp//24_04_17_00//cz_products_stock_updated.xlsx')


def set_img():
    url_money = 'C://Users//ww//Desktop//excel//eshp//24_04_17_00//cz_products_images.xlsx'
    wb_money = excel.open(url_money)
    ws_money = wb_money[wb_money.sheetnames[0]]

    url_eshop = 'C://Users//ww//Desktop//excel//eshp//24_04_17_00//en_products_images.xlsx'
    wb_eshop = excel.open(url_eshop)
    ws_eshop = wb_eshop[wb_eshop.sheetnames[0]]

    for e_shop_row in ws_eshop.rows:
        for money_row in ws_money.rows:
            mone_code = money_row[0].value
            e_sh_code = e_shop_row[0].value

            if mone_code == e_sh_code:
                e_shop_row[4].value = money_row[4].value
                e_shop_row[5].value = money_row[5].value
                e_shop_row[6].value = money_row[6].value
                e_shop_row[7].value = money_row[7].value
                e_shop_row[8].value = money_row[8].value
                e_shop_row[9].value = money_row[9].value
                e_shop_row[10].value = money_row[10].value
                e_shop_row[11].value = money_row[11].value
                e_shop_row[12].value = money_row[12].value
                e_shop_row[13].value = money_row[13].value
                e_shop_row[14].value = money_row[14].value
                e_shop_row[15].value = money_row[15].value
    wb_eshop.save('C://Users//ww//Desktop//excel//eshp//24_04_17_00//en_products_images_updated.xlsx')


def change_names():
    url_money = 'C://Users//ww//Desktop//excel//eshp//28_04//Artikl_28_04.xlsx'
    wb_money = excel.open(url_money)
    ws_money = wb_money[wb_money.sheetnames[0]]

    url_eshop = 'C://Users//ww//Desktop//excel//eshp//28_04//cz_products_vis_pr.xlsx'
    wb_eshop = excel.open(url_eshop)
    ws_eshop = wb_eshop[wb_eshop.sheetnames[0]]

    for e_shop_row in ws_eshop.rows:
        for money_row in ws_money.rows:

            money_code = money_row[4].value
            e_eshop_code = e_shop_row[0].value

            if money_code == e_eshop_code:
                brand_name = money_row[6].value
                money_name = money_row[5].value

                if brand_name is not None and brand_name != '':
                    money_name = money_name.replace(brand_name, '')
                    money_name = brand_name + ' ' + money_name
                # index of name from money
                e_shop_row[2].value = money_name
                e_shop_row[3].value = brand_name
    wb_eshop.save('C://Users//ww//Desktop//excel//eshp//28_04//cz_products_names.xlsx')


def set_prices():
    url_money = 'C://Users//ww//Desktop//excel//eshp//28_04//cz_prices_old.xlsx'
    wb_money = excel.open(url_money)
    ws_money = wb_money[wb_money.sheetnames[0]]

    url_eshop = 'C://Users//ww//Desktop//excel//eshp//28_04//en_prices_changed.xlsx'
    wb_eshop = excel.open(url_eshop)
    ws_eshop = wb_eshop[wb_eshop.sheetnames[0]]

    for e_shop_row in ws_eshop.rows:
        for money_row in ws_money.rows:

            money_code = money_row[0].value
            e_eshop_code = e_shop_row[0].value
            if money_code == e_eshop_code:
                money_row[3].value = e_shop_row[3].value
                money_row[4].value = e_shop_row[4].value
                money_row[5].value = e_shop_row[5].value
                money_row[6].value = e_shop_row[6].value
                money_row[7].value = e_shop_row[7].value

    wb_money.save('C://Users//ww//Desktop//excel//eshp//28_04//cz_prices_to_import.xlsx')


# set_stock()

# set_img()

# change_names()
set_prices()
